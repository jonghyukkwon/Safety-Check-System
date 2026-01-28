import streamlit as st
import google.generativeai as genai
import json
import io
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from guide_data import MASTER_GUIDE_TEXT

# ==========================================
# 1. API ì„¤ì • ë° ëª¨ë¸ ì„ ì–¸
# ==========================================
try:
    API_KEY = st.secrets["GEMINI_API_KEY"]
except:
    API_KEY = "YOUR_GEMINI_API_KEY" # ë¡œì»¬ í…ŒìŠ¤íŠ¸ìš© í‚¤

genai.configure(api_key=API_KEY)

generation_config = {
    "temperature": 0.0,
    "top_p": 1,
    "top_k": 1,
    "max_output_tokens": 8000,
}

creative_config = {
    "temperature": 0.2, # ìœ„í—˜ì„±í‰ê°€ ìƒì„±ì€ ì•½ê°„ì˜ ì°½ì˜ì„±ì´ í•„ìš”í•˜ë¯€ë¡œ 0.2ë¡œ ì„¤ì •
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8000,
}

MODEL_ID = "models/gemini-2.5-flash"

# ==========================================
# 2. ì—‘ì…€ ì–‘ì‹ ìƒì„± ë° ë°ì´í„° ì…ë ¥ í•¨ìˆ˜
# ==========================================
def generate_excel_from_scratch(p_info, risk_data):
    """
    ë¹ˆ ì—‘ì…€ì´ ì•„ë‹ˆë¼, ì½”ë“œë¡œ ìŠ¤íƒ€ì¼(í…Œë‘ë¦¬, ìƒ‰ìƒ)ì„ ì§ì ‘ ê·¸ë ¤ì„œ 
    ì™„ì„±ëœ í˜•íƒœì˜ ì—‘ì…€ íŒŒì¼ì„ ìƒì„±í•˜ëŠ” í•¨ìˆ˜
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ìœ„í—˜ì„±í‰ê°€ì„œ"

    # --- ìŠ¤íƒ€ì¼ ì •ì˜ ---
    # 1. í…Œë‘ë¦¬ ìŠ¤íƒ€ì¼ (ì–‡ì€ ì‹¤ì„ )
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # 2. í—¤ë” ìŠ¤íƒ€ì¼ (íšŒìƒ‰ ë°°ê²½, êµµì€ ê¸€ì”¨, ì¤‘ì•™ ì •ë ¬)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 3. ì œëª© ìŠ¤íƒ€ì¼
    title_font = Font(bold=True, size=16)

    # --- 1. ë¬¸ì„œ ì œëª© ì‘ì„± ---
    ws.merge_cells('B2:F2')
    ws['B2'] = "ê³µì‚¬ ë° ì‘ì—… ì•ˆì „ë³´ê±´ ìœ„í—˜ì„±í‰ê°€ì„œ"
    ws['B2'].font = title_font
    ws['B2'].alignment = center_align

    # --- 2. ê³µì‚¬ ê°œìš” (í‘œ ìƒë‹¨) ì‘ì„± ---
    # ë ˆì´ë¸” (Bì—´)
    labels = ["ê³µì‚¬ëª…", "ê³µì‚¬ ì¥ì†Œ", "ê³µì‚¬ ê¸°ê°„", "ì‘ì—… ë‚´ìš©"]
    keys = ["name", "loc", "period", "content"]
    
    start_row = 4
    for i, label in enumerate(labels):
        row = start_row + i
        # ë ˆì´ë¸” ì…€ (Bì—´)
        ws.cell(row=row, column=2, value=label).fill = header_fill
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border
        
        # ë°ì´í„° ì…€ (C~Fì—´ ë³‘í•©)
        ws.merge_cells(f'C{row}:F{row}')
        cell = ws.cell(row=row, column=3, value=p_info[keys[i]])
        cell.alignment = left_align
        cell.border = thin_border
        # ë³‘í•©ëœ ì…€ í…Œë‘ë¦¬ ì ìš©ì„ ìœ„í•œ ì²˜ë¦¬
        for col in range(3, 7):
            ws.cell(row=row, column=col).border = thin_border

    # --- 3. ìœ„í—˜ì„±í‰ê°€ í‘œ í—¤ë” ì‘ì„± ---
    table_header_row = start_row + 5 # ê°œìš” ë°‘ì— ë„ìš°ê³  ì‹œì‘
    headers = ["êµ¬ë¶„ (ì¥ë¹„/ì‘ì—…)", "ìœ„í—˜ìš”ì¸ (What)", "ìœ„í—˜ì„±", "ì•ˆì „ëŒ€ì±… (How)", "ë‹´ë‹¹ì"]
    col_widths = [20, 40, 10, 50, 15] # ì—´ ë„ˆë¹„ ì„¤ì •

    for i, header in enumerate(headers):
        col_idx = i + 2 # Bì—´(2)ë¶€í„° ì‹œì‘
        cell = ws.cell(row=table_header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # ì—´ ë„ˆë¹„ ì¡°ì •
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[i]

    # --- 4. AI ë°ì´í„° ì±„ìš°ê¸° ---
    current_row = table_header_row + 1
    
    for item in risk_data:
        # ë°ì´í„° ë§¤í•‘
        row_data = [
            item.get('equipment', ''),
            item.get('risk_factor', ''),
            item.get('risk_level', ''),
            item.get('countermeasure', ''),
            item.get('manager', '')
        ]
        
        for i, val in enumerate(row_data):
            col_idx = i + 2
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align if i != 3 else left_align # ëŒ€ì±…ë§Œ ì™¼ìª½ ì •ë ¬
            
            # ì¤„ë°”ê¿ˆ í—ˆìš© (ë‚´ìš©ì´ ê¸¸ ê²½ìš°)
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal, 
                                     vertical="center", 
                                     wrap_text=True)
            
        current_row += 1

    # --- 5. ê²°ì¬ë€ ë§Œë“¤ê¸° (ì„ íƒì‚¬í•­) ---
    sign_row = current_row + 2
    ws.merge_cells(f'B{sign_row}:F{sign_row}')
    ws[f'B{sign_row}'] = "ìœ„ì™€ ê°™ì´ ìœ„í—˜ì„±í‰ê°€ë¥¼ ì‹¤ì‹œí•˜ê³  ì•ˆì „ì¡°ì¹˜ë¥¼ ì´í–‰í•˜ê² ìŠµë‹ˆë‹¤."
    ws[f'B{sign_row}'].alignment = center_align
    
    sign_row += 2
    ws.cell(row=sign_row, column=4, value="ì‘ì„±ì(ì‹œê³µì‚¬): (ì¸)")
    ws.cell(row=sign_row, column=6, value="í™•ì¸ì(ê°ë…ì): (ì¸)")

    # íŒŒì¼ ì €ì¥ (ë©”ëª¨ë¦¬)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# 3. ë©”ì¸ UI êµ¬ì„±
# ==========================================
st.set_page_config(page_title="í˜¸í…” ì•ˆì „ë³´ê±´ ì‹œìŠ¤í…œ", layout="wide")

# ìƒë‹¨ íˆ´ë°” ë° ë°°ê²½ ìŠ¤íƒ€ì¼ ì»¤ìŠ¤í…€
st.markdown("""
    <style>
        /* ìƒë‹¨ í—¤ë” ì˜ì—­ ë°°ê²½ìƒ‰ ì„¤ì • */
        header[data-testid="stHeader"] {
            background-color: #9F896C !important;
        }
        
        /* í—¤ë” ë‚´ ì•„ì´ì½˜ ìƒ‰ìƒì„ í™”ì´íŠ¸ë¡œ ë³€ê²½ (ê³¨ë“œ ë°°ê²½ ëŒ€ë¹„) */
        header[data-testid="stHeader"] svg {
            fill: white !important;
        }

        /* íƒ­ ë°”ì˜ ê°•ì¡° ë¼ì¸ ìƒ‰ìƒ */
        .stTabs [data-baseweb="tab-highlight-indicator"] {
            background-color: #9F896C !important;
        }
        
        /* ë²„íŠ¼ ë°°ê²½ìƒ‰ ì»¤ìŠ¤í…€ */
        div.stButton > button:first-child {
            background-color: #9F896C;
            color: white;
            border: none;
        }
    </style>
    """, unsafe_allow_html=True)

st.title("ğŸ¨ í˜¸í…” ì•ˆì „ë³´ê±´ í†µí•© ê´€ë¦¬ ì‹œìŠ¤í…œ")

tab1, tab2, tab3 = st.tabs(["ğŸ“‘ ì ê²©ìˆ˜ê¸‰ì—…ì²´ í‰ê°€", "ğŸ“Š ìœ„í—˜ì„±í‰ê°€ ìë™ ìƒì„±", "ğŸ“‘ ì•ˆì „ë³´ê±´ê´€ë¦¬ê³„íšì„œ ê¸°ë°˜ ìœ„í—˜ì„±í‰ê°€ ìƒì„±"])

# --- TAB 1: ê¸°ì¡´ ì½”ë“œ (ìœ ì§€) ---

# ------------------------------------------------------------------------------
# TAB 1: ì ê²©ìˆ˜ê¸‰ì—…ì²´ í‰ê°€ (ë¬¸ë²• ì˜¤ë¥˜ ìˆ˜ì • ì™„ë£Œ)
# ------------------------------------------------------------------------------
with tab1:
    st.header("1. ìˆ˜ê¸‰ì—…ì²´ ì•ˆì „ë³´ê±´ê´€ë¦¬ê³„íšì„œ í‰ê°€")
    st.info("AIê°€ ê°€ì´ë“œë¼ì¸ì— ë”°ë¼ ì ìˆ˜ë¥¼ ì‚°ì¶œí•©ë‹ˆë‹¤.")

    # 1. ëª¨ë¸ ì„¤ì •
    eval_model = genai.GenerativeModel(
        model_name="models/gemini-2.5-flash",
        generation_config={
            "temperature": 0.0,
            "response_mime_type": "application/json",
        },
        system_instruction=  "ë‹¹ì‹ ì€ ì°½ì˜ì„±ì´ ì—†ëŠ” 'ì•ˆì „ë³´ê±´ ì ìˆ˜ ê³„ì‚°ê¸°'ì…ë‹ˆë‹¤. "

            "ë¬¸ì„œë¥¼ í•´ì„í•˜ë ¤ í•˜ì§€ ë§ê³ , í…ìŠ¤íŠ¸ì— í‚¤ì›Œë“œê°€ ìˆëŠ”ì§€ë§Œ í™•ì¸í•˜ì‹­ì‹œì˜¤. "

            "ë‹¤ìŒ 'ì±„ì  ì•Œê³ ë¦¬ì¦˜'ì„ ì—„ê²©íˆ ë”°ë¥´ì‹­ì‹œì˜¤: "

            "1. IF (í•´ë‹¹ í•­ëª©ì— ëŒ€í•œ êµ¬ì²´ì  ê³„íš + ì‹¤í–‰ ì¦ë¹™/ì‚¬ì§„) EXIST -> THEN [ìš°ìˆ˜/ë§Œì ] "

            "2. IF (ê³„íšì€ ìˆìœ¼ë‚˜ ì¦ë¹™ì´ ì—†ê±°ë‚˜ ì¶”ìƒì ì„) -> THEN [ë³´í†µ/ì¤‘ê°„ì ìˆ˜] "

            "3. IF (ë‚´ìš© ëˆ„ë½ OR ë‹¤ë¥¸ í˜„ì¥ ë³µì‚¬/ë¶™ì—¬ë„£ê¸° í”ì ) -> THEN [ë¯¸í¡/ìµœí•˜ì ] "

            "4. DEFAULT (íŒë‹¨ì´ ì• ë§¤í•œ ëª¨ë“  ê²½ìš°) -> THEN [ë¯¸í¡/ìµœí•˜ì ] (ì ìˆ˜ ë¶€í’€ë¦¬ê¸° ê¸ˆì§€)"
    )

    user_file = st.file_uploader("ì—…ì²´ ì œì¶œ ê³„íšì„œ(PDF) ì—…ë¡œë“œ", type=["pdf"], key="eval_upload")

    if st.button("í‰ê°€ ì‹œì‘", key="eval_btn"):
        if not user_file:
            st.warning("íŒŒì¼ì„ ì—…ë¡œë“œí•´ ì£¼ì„¸ìš”.")
        else:
            with st.spinner("AIê°€ ë¬¸ì„œì˜ ì´ë¯¸ì§€ì™€ ë‚´ìš©ì„ ì •ë°€ ë¶„ì„ ì¤‘..."):
                temp_path = "temp_eval.pdf"
                try:
                    # íŒŒì¼ ì €ì¥ ë° ì—…ë¡œë“œ
                    with open(temp_path, "wb") as f:
                        f.write(user_file.getbuffer())
                    
                    uploaded_file = genai.upload_file(temp_path, mime_type="application/pdf")
                    
                    # íŒŒì¼ ì²˜ë¦¬ ëŒ€ê¸°
                    while uploaded_file.state.name == "PROCESSING":
                        time.sleep(1)
                        uploaded_file = genai.get_file(uploaded_file.name)

                    # í”„ë¡¬í”„íŠ¸
                    prompt = f"""
                  
                    [ì°¸ì¡°: ê°€ì´ë“œë¼ì¸]
                    {MASTER_GUIDE_TEXT}

                    [ë§ˆìŠ¤í„° ê°€ì´ë“œë¼ì¸]ì„ ê¸°ì¤€ìœ¼ë¡œ ìˆ˜ê¸‰ì—…ì²´ ê³„íšì„œë¥¼ ì±„ì í•˜ì‹­ì‹œì˜¤.

                    ë³€ë•ìŠ¤ëŸ¬ìš´ ì ìˆ˜ë¥¼ ë§‰ê¸° ìœ„í•´, ê° í•­ëª©ë³„ë¡œ **ë°˜ë“œì‹œ PDF ë‚´ì˜ 'ì¦ê±° ë¬¸ì¥'ì„ ë¨¼ì € ì°¾ê³ ** ì ìˆ˜ë¥¼ ë§¤ê¸°ì‹­ì‹œì˜¤.

                    
                [ğŸš« ì ˆëŒ€ì  ì±„ì  ê·œì¹™ (Tie-Breaker Rule)]

                    1. **ì¦ê±° ìš°ì„ ì£¼ì˜**: "ì˜ í•  ê²ƒìœ¼ë¡œ ë³´ì„", "ê³„íšëœ ê²ƒìœ¼ë¡œ ì¶”ì •ë¨" ê°™ì€ ì¶”ì¸¡ì€ ì ˆëŒ€ ê¸ˆì§€. PDFì— ëª…ì‹œëœ ë¬¸êµ¬ê°€ ì—†ìœ¼ë©´ 0ì .

                    2. **í•˜í–¥ í‰ê°€ ì›ì¹™**: 

                       - 5ì  ì¤„ê¹Œ 3ì  ì¤„ê¹Œ ê³ ë¯¼ë˜ë©´ -> **3ì ** ë¶€ì—¬.

                       - 3ì  ì¤„ê¹Œ 1ì  ì¤„ê¹Œ ê³ ë¯¼ë˜ë©´ -> **1ì ** ë¶€ì—¬.

                       - **ì¦‰, í™•ì‹¤í•œ ê·¼ê±°ê°€ ì—†ëŠ” í•œ ë†’ì€ ì ìˆ˜ë¥¼ ì£¼ì§€ ë§ˆì‹œì˜¤.**

                    3. **ê³µì¢… ì¼ì¹˜ì„±**: PDF ì œëª©ì˜ ê³µì‚¬ëª…ê³¼ ë³¸ë¬¸ì˜ ì‘ì—… ë‚´ìš©ì´ ë¶ˆì¼ì¹˜(ë³µì‚¬ ë¶™ì—¬ë„£ê¸° ì˜ì‹¬)í•˜ë©´ í•´ë‹¹ í•­ëª© 0ì  ì²˜ë¦¬.

                    4. **ì¤‘ëŒ€ì¬í•´(17ë²ˆ)**: 'í•´ë‹¹ì—†ìŒ' ë˜ëŠ” 'ë¬´ì¬í•´'ë¼ëŠ” ëª…í™•í•œ í…ìŠ¤íŠ¸ë‚˜ ì¦ëª…ì„œê°€ ì—†ìœ¼ë©´, í™•ì¸ ë¶ˆê°€ë¡œ ê°„ì£¼í•˜ì—¬ 0ì  ì²˜ë¦¬.

                    [ì¶œë ¥ í˜•ì‹]
                    [
                        {{
                            "item_no": 1,
                            "category": "í•­ëª©ëª…",
                            "score": 0,
                            "max_score": 5,
                            "evidence": "ì¦ê±° ë‚´ìš©",
                            "judgment": "ë“±ê¸‰"
                        }}
                          ... (17ë²ˆê¹Œì§€ ë°˜ë³µ)
                    ]
                    """
                    
                    # AI ì‹¤í–‰
                    response = eval_model.generate_content([prompt, uploaded_file])
                    
                    # ê²°ê³¼ ì²˜ë¦¬
                    eval_data = json.loads(response.text)
                    
                    # ë°ì´í„° ìœ íš¨ì„± ê²€ì‚¬ (ë¦¬ìŠ¤íŠ¸ì¸ì§€ í™•ì¸)
                    if isinstance(eval_data, list):
                        total_score = sum(item['score'] for item in eval_data)
                        
                        st.markdown(f"## ğŸ† ì¢…í•© ì ìˆ˜: **{total_score}ì **")
                        st.markdown("---")
                        
                        # ìë™ ì´í‰ ì¶œë ¥
                        if total_score >= 90:
                            st.success("âœ… **[ê³ ìœ„í—˜êµ° / ì¼ë°˜êµ° ëª¨ë‘ ì ê²©]**")
                        elif 80 <= total_score < 90:
                            st.warning("âš ï¸ **[ì¼ë°˜êµ° ì ê²© / ê³ ìœ„í—˜êµ° ë¶€ì ê²©]**")
                        elif 70 <= total_score < 80:
                            st.error("âŒ **[ë¶€ì ê²©]** (80ì  ë¯¸ë‹¬)")
                        else:
                            st.error("ğŸš« **[ì ˆëŒ€ ì„ ì • ë¶ˆê°€]** (70ì  ë¯¸ë§Œ)")
                        
                        st.markdown("---")
                        
                        # í…Œì´ë¸” ë³€í™˜ ë° ì¶œë ¥
                        display_data = []
                        for item in eval_data:
                            display_data.append({
                                "í•­ëª©": f"{item['item_no']}. {item['category']}",
                                "ì ìˆ˜": f"{item['score']} / {item['max_score']}",
                                "ë“±ê¸‰": item['judgment'],
                                "ê·¼ê±°": item['evidence']
                            })
                        st.table(display_data)

                    else:
                        # ë¦¬ìŠ¤íŠ¸ê°€ ì•„ë‹ ê²½ìš° ì—ëŸ¬ ì²˜ë¦¬
                        st.error("AIê°€ ì˜¬ë°”ë¥¸ ë¦¬ìŠ¤íŠ¸ í˜•ì‹ì˜ ë°ì´í„°ë¥¼ ë°˜í™˜í•˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.")
                        st.json(eval_data) # ë””ë²„ê¹…ìš© ë°ì´í„° ì¶œë ¥

                    # ë’·ì •ë¦¬
                    genai.delete_file(uploaded_file.name)
                    if os.path.exists(temp_path): os.remove(temp_path)

                except Exception as e:
                    st.error(f"ì˜¤ë¥˜ ë°œìƒ: {e}")
                    if os.path.exists(temp_path): os.remove(temp_path)
                        
# --- TAB 2: ì—‘ì…€ ìë™ ìƒì„± (NEW) ---
with tab2:
    st.header("2. ê³µì‚¬ ìœ„í—˜ì„±í‰ê°€ ì—‘ì…€(Excel) ìë™ ì‘ì„±")
    st.info("ê³µì‚¬ ë‚´ìš©ì„ ì…ë ¥í•˜ë©´, AIê°€ **í‘œì¤€ ì„œì‹ì´ ì ìš©ëœ ì—‘ì…€ íŒŒì¼**ì„ ì¦‰ì‹œ ë§Œë“¤ì–´ì¤ë‹ˆë‹¤.")

    with st.container(border=True):
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("ğŸ“ 1. ê³µì‚¬ ê°œìš” ì…ë ¥")
            p_name = st.text_input("ê³µì‚¬ëª…", placeholder="ì˜ˆ: 3ì¸µ ê°ì‹¤ ë¦¬ëª¨ë¸ë§")
            p_loc = st.text_input("ì¥ì†Œ", placeholder="ì˜ˆ: ë³¸ê´€ 3ì¸µ")
            p_period = st.text_input("ê¸°ê°„", placeholder="ì˜ˆ: 26.02.01 ~ 02.15")
            p_content = st.text_area("ì‘ì—… ìƒì„¸ ë‚´ìš©", height=120, placeholder="êµ¬ì²´ì ì¸ ì‘ì—… ë‚´ìš© ì…ë ¥...")

        with col2:
            st.subheader("âš ï¸ 2. ìœ„í—˜ìš”ì¸ ì„ íƒ")
            risk_cols = st.columns(3)
            r_check = [
                risk_cols[0].checkbox("ğŸ”¥ í™”ê¸°"),
                risk_cols[0].checkbox("âš¡ ì „ê¸°"),
                risk_cols[1].checkbox("ğŸªœ ê³ ì†Œ"),
                risk_cols[1].checkbox("ğŸ—ï¸ ì¤‘ëŸ‰ë¬¼"),
                risk_cols[2].checkbox("â˜ ï¸ ìœ„í—˜ë¬¼"),
                risk_cols[2].checkbox("ğŸ•³ï¸ ë°€í")
            ]
            risk_labels = ["í™”ê¸°", "ì „ê¸°", "ê³ ì†Œ", "ì¤‘ëŸ‰ë¬¼", "ìœ„í—˜ë¬¼", "ë°€í"]
            selected_risks = [label for label, checked in zip(risk_labels, r_check) if checked]

            st.markdown("---")
            generate_btn = st.button("âœ¨ ì—‘ì…€ íŒŒì¼ ìƒì„±í•˜ê¸° (AI)", type="primary", use_container_width=True)

    if generate_btn:
        if not p_name or not selected_risks:
            st.warning("ê³µì‚¬ëª…ê³¼ ìµœì†Œ 1ê°œì˜ ìœ„í—˜ìš”ì¸ì„ ì„ íƒí•´ì£¼ì„¸ìš”.")
        else:
            with st.spinner("AIê°€ ìœ„í—˜ìš”ì¸ì„ ë¶„ì„í•˜ê³  ì—‘ì…€ ì„œì‹ì„ ê·¸ë¦¬ëŠ” ì¤‘ì…ë‹ˆë‹¤..."):
                try:
                    # 1. ëª¨ë¸ í˜¸ì¶œ ë° ë°ì´í„° ìƒì„±
                    model = genai.GenerativeModel(MODEL_ID)
                    
                    prompt = f"""
                    ë‹¤ìŒ ê³µì‚¬ ì •ë³´ë¥¼ ë°”íƒ•ìœ¼ë¡œ [ìœ„í—˜ì„±í‰ê°€í‘œ]ì— ë“¤ì–´ê°ˆ ë‚´ìš©ì„ JSON í˜•ì‹ìœ¼ë¡œ ì‘ì„±í•´ì¤˜.
                    
                    [ê³µì‚¬ ì •ë³´]
                    - ê³µì‚¬ëª…: {p_name} / ë‚´ìš©: {p_content}
                    - í•µì‹¬ ìœ„í—˜ìš”ì¸: {", ".join(selected_risks)}

                    [ì‘ì„± ê·œì¹™]
                    1. ì„ íƒëœ ìœ„í—˜ìš”ì¸ê³¼ ê´€ë ¨ëœ êµ¬ì²´ì ì¸ ìœ„í—˜ í•­ëª©ì„ 5~7ê°œ ë„ì¶œí•  ê²ƒ.
                    2. ê°ì†ŒëŒ€ì±…ì€ "ì•ˆì „ëª¨ ì°©ìš©" ì²˜ëŸ¼ ì§§ê²Œ ì“°ì§€ ë§ê³ , "KCS ì¸ì¦ ì•ˆì „ëª¨ ì°©ìš© ë° í„±ëˆ ì²´ê²° í™•ì¸" ì²˜ëŸ¼ êµ¬ì²´ì ìœ¼ë¡œ ì‘ì„±í•  ê²ƒ.
                    3. **ë°˜ë“œì‹œ ì•„ë˜ JSON êµ¬ì¡°ë§Œ ì¶œë ¥í•  ê²ƒ.**
                    
                    [
                        {{
                            "equipment": "ì‘ì—…ëª…/ì¥ë¹„ (ì˜ˆ: ìš©ì ‘ì‘ì—…)",
                            "risk_factor": "ìœ„í—˜ìš”ì¸ (ì˜ˆ: ë¶ˆí‹° ë¹„ì‚°)",
                            "risk_level": "ìƒ/ì¤‘/í•˜",
                            "countermeasure": "êµ¬ì²´ì  ëŒ€ì±… ë‚´ìš©ì„ ê¸¸ê²Œ ì‘ì„±",
                            "manager": "ì•ˆì „ë‹´ë‹¹ì"
                        }}
                    ]
                    """
                    
                    response = model.generate_content(prompt)
                    
                    # 2. JSON íŒŒì‹±
                    clean_text = re.sub(r"```json|```", "", response.text).strip()
                    risk_data_list = json.loads(clean_text)
                    
                    # 3. ì—‘ì…€ íŒŒì¼ ìƒì„± (ìŠ¤íƒ€ì¼ ì ìš©)
                    p_info = {"name": p_name, "loc": p_loc, "period": p_period, "content": p_content}
                    excel_byte = generate_excel_from_scratch(p_info, risk_data_list)
                    
                    # 4. ë‹¤ìš´ë¡œë“œ ì œê³µ
                    st.success("âœ… ì—‘ì…€ íŒŒì¼ ìƒì„±ì´ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤!")
                    st.download_button(
                        label="ğŸ“¥ ì—‘ì…€ íŒŒì¼ ë‹¤ìš´ë¡œë“œ (.xlsx)",
                        data=excel_byte,
                        file_name=f"ìœ„í—˜ì„±í‰ê°€_{p_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except json.JSONDecodeError:
                    st.error("AI ì‘ë‹µ ì²˜ë¦¬ ì‹¤íŒ¨ (ë°ì´í„° í˜•ì‹ ì˜¤ë¥˜). ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”.")
                except Exception as e:

                    st.error(f"ì˜¤ë¥˜ ë°œìƒ: {e}")



# TAB 3: ìœ„í—˜ì„±í‰ê°€ ì—‘ì…€ ìƒì„± (PDF ê¸°ë°˜ - NEW)
# ------------------------------------------------------------------------------
with tab3:
    st.header("3. ì•ˆì „ë³´ê±´ê´€ë¦¬ê³„íšì„œ ê¸°ë°˜ ìœ„í—˜ì„±í‰ê°€ ìë™ ìƒì„±")
    st.info("ì—…ë¡œë“œí•œ ê³„íšì„œì˜ **ê³µì‚¬ ê°œìš”**ì™€ **ì‘ì—… ë‚´ìš©**ì„ AIê°€ ìŠ¤ìŠ¤ë¡œ ë¶„ì„í•˜ì—¬, ìœ„í—˜ì„±í‰ê°€ ì—‘ì…€ íŒŒì¼ì„ ë§Œë“¤ì–´ì¤ë‹ˆë‹¤.")

    pdf_file = st.file_uploader("ì•ˆì „ë³´ê±´ê´€ë¦¬ê³„íšì„œ(PDF) ì—…ë¡œë“œ", type=["pdf"], key="risk_pdf_upload")
    
    if st.button("ğŸš€ PDF ë¶„ì„ ë° ì—‘ì…€ ìƒì„±", key="pdf_risk_btn", type="primary"):
        if not pdf_file:
            st.warning("ë¨¼ì € PDF íŒŒì¼ì„ ì—…ë¡œë“œí•´ì£¼ì„¸ìš”.")
        else:
            with st.spinner("AIê°€ ê³„íšì„œë¥¼ ë¶„ì„í•˜ì—¬ ê³µì‚¬ ì •ë³´ì™€ ìœ„í—˜ìš”ì¸ì„ ì¶”ì¶œí•˜ê³  ìˆìŠµë‹ˆë‹¤..."):
                temp_pdf_path = "temp_plan.pdf"
                try:
                    # 1. PDF ì—…ë¡œë“œ
                    with open(temp_pdf_path, "wb") as f:
                        f.write(pdf_file.getbuffer())
                    
                    uploaded_pdf = genai.upload_file(temp_pdf_path, mime_type="application/pdf")
                    while uploaded_pdf.state.name == "PROCESSING":
                        time.sleep(1)
                        uploaded_pdf = genai.get_file(uploaded_pdf.name)

                    # 2. AI ë¶„ì„ ëª¨ë¸ í˜¸ì¶œ
                    # (ì°½ì˜ì„± ì„¤ì • ì ìš©: ìœ„í—˜ìš”ì¸ ë„ì¶œ ì‹œ ìœ ì—°ì„± í•„ìš”)
                    pdf_risk_model = genai.GenerativeModel(MODEL_ID, generation_config=creative_config)

                    prompt = """
                    ì²¨ë¶€ëœ [ì•ˆì „ë³´ê±´ê´€ë¦¬ê³„íšì„œ] PDFë¥¼ ì •ë°€ ë¶„ì„í•˜ì—¬, ì•„ë˜ ë‘ ê°€ì§€ ì •ë³´ë¥¼ JSON í˜•ì‹ìœ¼ë¡œ ì¶”ì¶œ ë° ìƒì„±í•˜ì„¸ìš”.

                    1. **ê³µì‚¬ ê°œìš” ì •ë³´ ì¶”ì¶œ**: PDF ë‚´ì—ì„œ ê³µì‚¬ëª…, í˜„ì¥ ìœ„ì¹˜(ì¥ì†Œ), ê³µì‚¬ ê¸°ê°„, ì£¼ìš” ì‘ì—… ë‚´ìš©ì„ ì°¾ì•„ë‚´ì„¸ìš”.
                       (ë§Œì•½ ì •í™•í•œ ê¸°ê°„ì´ë‚˜ ì¥ì†Œê°€ ëª…ì‹œë˜ì§€ ì•Šì•˜ë‹¤ë©´ 'PDF ë‚´ ë¯¸ê¸°ì¬'ë¡œ í‘œê¸°í•  ê²ƒ)
                    
                    2. **ìœ„í—˜ì„±í‰ê°€ ë°ì´í„° ìƒì„±**: 
                       - ë¶„ì„ëœ 'ì‘ì—… ë‚´ìš©'ê³¼ 'í˜„ì¥ ì‚¬ì§„/ë„ë©´' ë“±ì„ ë°”íƒ•ìœ¼ë¡œ ì˜ˆìƒë˜ëŠ” ì£¼ìš” ìœ„í—˜ìš”ì¸ì„ 7ê°€ì§€ ì´ìƒ ë„ì¶œí•˜ì„¸ìš”.
                       - ê° ìœ„í—˜ìš”ì¸ì— ëŒ€í•´ êµ¬ì²´ì ì¸ ì•ˆì „ ëŒ€ì±…ì„ ìˆ˜ë¦½í•˜ì„¸ìš”.

                    [í•„ìˆ˜ ì¶œë ¥ í˜•ì‹ (JSON Only)]
                    ë°˜ë“œì‹œ ì•„ë˜ JSON êµ¬ì¡°ë¥¼ ì—„ê²©íˆ ì§€ì¼œì„œ ì¶œë ¥í•˜ì„¸ìš”. Markdown ì½”ë“œëŠ” ì œì™¸í•˜ì„¸ìš”.

                    {
                        "project_info": {
                            "name": "ê³µì‚¬ëª… ì¶”ì¶œ ê²°ê³¼",
                            "loc": "ì¥ì†Œ ì¶”ì¶œ ê²°ê³¼",
                            "period": "ê¸°ê°„ ì¶”ì¶œ ê²°ê³¼",
                            "content": "ì‘ì—… ë‚´ìš© ìš”ì•½"
                        },
                        "risk_data": [
                            {
                                "equipment": "ì‘ì—…ë‹¨ìœ„ ë˜ëŠ” ì‚¬ìš©ì¥ë¹„ (ì˜ˆ: ìš©ì ‘ì‘ì—…)",
                                "risk_factor": "êµ¬ì²´ì  ìœ„í—˜ìš”ì¸ (ì˜ˆ: ë¶ˆí‹° ë¹„ì‚°ì— ì˜í•œ í™”ì¬)",
                                "risk_level": "ìƒ/ì¤‘/í•˜",
                                "countermeasure": "êµ¬ì²´ì  ì•ˆì „ ëŒ€ì±… (KCS ì¸ì¦ ë³´í˜¸êµ¬ ì°©ìš© ë“±)",
                                "manager": "ì•ˆì „ë‹´ë‹¹ì"
                            }
                        ]
                    }
                    """

                    response = pdf_risk_model.generate_content([prompt, uploaded_pdf])
                    
                    # 3. ë°ì´í„° íŒŒì‹±
                    raw_text = response.text
                    # JSON ë¸”ë¡ ì°¾ê¸° ({ ... })
                    json_match = re.search(r'\{.*\}', raw_text, re.DOTALL)
                    
                    if json_match:
                        full_data = json.loads(json_match.group(0))
                        
                        # ë°ì´í„° ë¶„ë¦¬
                        extracted_info = full_data.get("project_info", {})
                        extracted_risks = full_data.get("risk_data", [])

                        # 4. ë¶„ì„ ê²°ê³¼ ë¯¸ë¦¬ë³´ê¸° (ì‚¬ìš©ì í™•ì¸ìš©)
                        st.success("âœ… ë¶„ì„ ì™„ë£Œ! ì¶”ì¶œëœ ì •ë³´ë¥¼ í™•ì¸í•˜ì„¸ìš”.")
                        
                        with st.expander("ğŸ“„ ì¶”ì¶œëœ ê³µì‚¬ ê°œìš” í™•ì¸", expanded=True):
                            c1, c2 = st.columns(2)
                            c1.text_input("ê³µì‚¬ëª…", value=extracted_info.get("name", ""), disabled=True)
                            c1.text_input("ì¥ì†Œ", value=extracted_info.get("loc", ""), disabled=True)
                            c2.text_input("ê¸°ê°„", value=extracted_info.get("period", ""), disabled=True)
                            c2.text_area("ì‘ì—… ë‚´ìš©", value=extracted_info.get("content", ""), disabled=True)

                        # 5. ì—‘ì…€ ìƒì„±
                        excel_byte = generate_excel_from_scratch(extracted_info, extracted_risks)

                        st.markdown("---")
                        st.download_button(
                            label="ğŸ“¥ ìœ„í—˜ì„±í‰ê°€ ì—‘ì…€ ë‹¤ìš´ë¡œë“œ (.xlsx)",
                            data=excel_byte,
                            file_name=f"ìœ„í—˜ì„±í‰ê°€_{extracted_info.get('name', 'ìë™ìƒì„±')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                        
                    else:
                        st.error("AI ì‘ë‹µì—ì„œ ìœ íš¨í•œ JSON ë°ì´í„°ë¥¼ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")
                        st.text(raw_text)

                    # íŒŒì¼ ì •ë¦¬
                    genai.delete_file(uploaded_pdf.name)
                    if os.path.exists(temp_pdf_path): os.remove(temp_pdf_path)

                except Exception as e:
                    st.error(f"ë¶„ì„ ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {e}")
                    if os.path.exists(temp_pdf_path): os.remove(temp_pdf_path)





















